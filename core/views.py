import random
from django.http import response
from django.shortcuts import render, redirect
from django.views import View
from django.utils import timezone
from django.contrib.auth.forms import UserCreationForm
from django.views.generic import CreateView, ListView, DeleteView, TemplateView
from django.views.generic.edit import UpdateView
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http.response import HttpResponse
from django.urls import reverse_lazy
from django.core.mail import send_mail
from django import forms
from django_filters.views import FilterView
from django.views.defaults import page_not_found
from openpyxl import Workbook

from blog.models import BlogEntry, BlogCategory
from contac.models import Contact
from audit.signals import Audits
from newportal.settings.base import EMAIL_HOST_USER
from .models import Profile, Description, OtherSites, Quote, Suscriptor, FriendSites, Music

from .forms import ProfileForm, EmailForm, NameUpdateForm, DescriptionForm, OtherSitesForm, QuoteForm, FriendSitesForm, SuscriptorEmailForm
from .filters import OtherSitesListFilter, QuoteListFilter, SuscriptorListFilter, FriendSitesFilter


class IndexView(View):
    def get(self, request, *args, **kwargs):
        try:
            posts = BlogEntry.objects.filter(
                active=True).order_by('-created_date')[:10]
            categories = BlogCategory.objects.filter(is_active=True)
            web = Description.objects.filter(is_active=True)
            #category = BlogCategory.objects.get(slug=slug)
            featured = BlogEntry.objects.filter(
                active=True, featured=True).order_by('-created_date')
            recientes = BlogEntry.objects.filter(
                active=True).order_by('-created_date')[0]
            sites = OtherSites.objects.filter(
                active=True).order_by('name')
            quotes = Quote.objects.filter(active=True)
            friendsites = FriendSites.objects.filter(
                active=True).order_by('-date')[:5]
            music = Music.objects.filter(is_active=True)

            context = {
                'posts': posts,

                'categories': categories,
                # 'category': category,
                'featured': featured,
                'friendsites': friendsites,
                'quote': random.choice(quotes),
                'sites': sites,
                'recientes': recientes,
                'music': music,
                'web': web,
            }
        except:
            context = {}
        return render(request, 'index.html', context)


class LateralView(ListView):
    def get(self, request, *args, **kwargs):
        try:
            featured = BlogEntry.objects.filter(
                active=True, featured=True).order_by('-created_date')
            sites = OtherSites.objects.filter(
                active=True).order_by('name')
            context = {
                'sites': sites,
                'featured': featured
            }
        except:
            context = {}
        return render(request, 'includes/lateral.html', context)


class NavbarView(View):
    def get(self, request, *args, **kwargs):
        try:
            categories = BlogCategory.objects.filter(is_active=True)
            context = {
                'categories': categories,
            }
        except:
            context = {}
        return render(request, 'templates/includes/navbar.html', context)


class SearchView(View):
    '''
    def get(self, request, *args, **kwargs):
        queryset = request.POST.get("buscar")
        categories = BlogCategory.objects.filter(is_active=True)
        web = Description.objects.filter(is_active=True)
        featured = BlogEntry.objects.filter(
            active=True, featured=True).order_by('-created_date')
        sites = OtherSites.objects.filter(
            active=True).order_by('name')
        context = {
            'categories': categories,
            'featured': featured,
            'sites': sites,
            'web': web,
        }

        return render(request, 'results.html', context)
        '''

    def post(self, request, *args, **kwargs):
        queryset = request.POST.get("buscar")
        categories = BlogCategory.objects.filter(is_active=True)
        web = Description.objects.filter(is_active=True)
        featured = BlogEntry.objects.filter(
            active=True, featured=True).order_by('-created_date')
        sites = OtherSites.objects.filter(
            active=True).order_by('name')
        quotes = Quote.objects.filter(active=True)
        if queryset:
            posts = BlogEntry.objects.filter(
                Q(name__icontains=queryset) |
                Q(description__icontains=queryset),
                active=True
            ).distinct().order_by('-created_date')

        context = {
            'categories': categories,
            'featured': featured,
            'quote': random.choice(quotes),
            'web': web,
            'sites': sites,
            'object_list': posts,
        }

        return render(request, 'results.html', context)


@method_decorator(login_required, name='dispatch')
class AdministrationView(View):
    def get(self, request, *args, **kwargs):
        total = len(list(Contact.objects.filter(state=True)))
        context = {
            'total': total
        }
        return render(request, "administration.html", context)


@method_decorator(login_required, name='dispatch')
# Actualiza o crea los datos de un usuario para un perfil
class ProfileUpdate(UpdateView):
    form_class = ProfileForm
    success_url = reverse_lazy('profile')
    template_name = 'registration/profile_form.html'

    def get_object(self):
        profile, created = Profile.objects.get_or_create(
            user=self.request.user)
        return profile


@method_decorator(login_required, name='dispatch')
class EmailUpdate(UpdateView):  # Permite editar el e-mail de un usuario
    form_class = EmailForm
    success_url = reverse_lazy('profile')
    template_name = 'registration/profile_email_form.html'

    def get_object(self):
        return self.request.user

    def get_form(self, form_class=None):
        form = super(EmailUpdate, self).get_form()
        # Modificar en tiempo real
        form.fields['email'].widget = forms.EmailInput(
            attrs={'class': 'form-control mb-2', 'placeholder': 'Email'})
        return form


@method_decorator(login_required, name='dispatch')
class AboutUsFormView(UpdateView):
    model = Description
    form_class = DescriptionForm
    success_url = reverse_lazy('administration')
    template_name = 'registration/about_us_form.html'


@method_decorator(login_required, name='dispatch')
class NameUpdate(UpdateView):  # Permite editar el e-mail de un usuario
    form_class = NameUpdateForm
    success_url = reverse_lazy('profile')
    template_name = 'registration/profile_name_form.html'

    def get_object(self):
        return self.request.user

    def get_form(self, form_class=None):
        form = super(NameUpdate, self).get_form()
        # Modificar en tiempo real
        form.fields['first_name'].widget = forms.TextInput(
            attrs={'class': 'form-control mb-2', 'placeholder': 'Nombre'})
        form.fields['last_name'].widget = forms.TextInput(
            attrs={'class': 'form-control mb-2', 'placeholder': 'Apellido'})
        return form


@method_decorator(login_required, name='dispatch')
class OtherSitesCreateView(CreateView):
    # Vista para crear los post con un de decorador para que solo los administardores puedan acceder
    model = OtherSites
    form_class = OtherSitesForm
    success_url = reverse_lazy('administration')
    template_name = 'othersides/othersites_form.html'

    # Devuelve al usuario actualmente logueado para el campo de User del Blog creado
    def form_valid(self, form):
        form.instance.user = self.request.user
        return super(OtherSitesCreateView, self).form_valid(form)


@method_decorator(login_required, name='dispatch')
class OtherSitesAdminListView(FilterView):
    # Vista de la lista de posts para el administrador
    model = OtherSites
    context_object_name = 'sites'
    template_name = 'othersides/adminothersites.html'
    paginate_by = 30  # TODO obtener dato de un constants.py
    filterset_class = OtherSitesListFilter
    ordering = ["name"]

    def get_queryset(self):
        return OtherSites.objects.all()


@method_decorator(login_required, name='dispatch')
class OtherSitesUpdateView(UpdateView):
    model = OtherSites
    form_class = OtherSitesForm
    template_name = 'othersides/othersites_update_form.html'
    template_name_suffix = '_update_form'

    def get_success_url(self):
        return reverse_lazy('sitesupdate', args=[self.object.id]) + '?ok'


@method_decorator(login_required, name='dispatch')
class OtherSitesChageStateView(View):
    # Cambia el estado de un blog, usado en el jquerry de blogadmin-list.html
    def post(self, request, site_id, *args, **kwargs):
        try:
            site = OtherSites.objects.get(id=site_id)
            site.active = not site.active
            site.save()
            if site.active:
                audit_description = f"Se activa el Sitio :{site_id}"
            else:
                audit_description = f"Se desactiva el Sitio :{site_id}"
            Audits.audit_action(request.user, "", audit_description, "DELETE")
            return HttpResponse("ok", status=200)
        except OtherSites.DoesNotExist:
            return HttpResponse("Sitio not found", status=404)
        except Exception as e:
            print(e)
            return HttpResponse("error code", status=500)


@method_decorator(login_required, name='dispatch')
class QuoteCreateView(CreateView):
    # Vista para crear los post con un de decorador para que solo los administardores puedan acceder
    model = Quote
    form_class = QuoteForm
    success_url = reverse_lazy('administration')
    template_name = 'quotes/quotes_form.html'

    # Devuelve al usuario actualmente logueado para el campo de User del Blog creado
    def form_valid(self, form):
        form.instance.user = self.request.user
        return super(QuoteCreateView, self).form_valid(form)


@method_decorator(login_required, name='dispatch')
class QuoteAdminListView(FilterView):
    # Vista de la lista de posts para el administrador
    model = Quote
    context_object_name = 'quotes'
    template_name = 'quotes/adminquotes.html'
    paginate_by = 30  # TODO obtener dato de un constants.py
    filterset_class = QuoteListFilter
    ordering = ["autor"]

    def get_queryset(self):
        return Quote.objects.all()


@method_decorator(login_required, name='dispatch')
class QuoteUpdateView(UpdateView):
    model = Quote
    form_class = QuoteForm
    template_name = 'quotes/quotes_update_form.html'
    template_name_suffix = '_update_form'

    def get_success_url(self):
        return reverse_lazy('quoteUpdate', args=[self.object.id]) + '?ok'


@method_decorator(login_required, name='dispatch')
class QuoteChageStateView(View):
    # Cambia el estado de un blog, usado en el jquerry de blogadmin-list.html
    def post(self, request, quote_id, *args, **kwargs):
        try:
            quote = Quote.objects.get(id=quote_id)
            quote.active = not quote.active
            quote.save()
            if quote.active:
                audit_description = f"Se activa el Sitio :{quote_id}"
            else:
                audit_description = f"Se desactiva el Sitio :{quote_id}"
            Audits.audit_action(request.user, "", audit_description, "DELETE")
            return HttpResponse("ok", status=200)
        except Quote.DoesNotExist:
            return HttpResponse("Cita not found", status=404)
        except Exception as e:
            print(e)
            return HttpResponse("error code", status=500)


class SuscribeView(View):
    def post(self, request, *args, **kwargs):
        email = request.POST.get('email')
        Suscriptor.objects.create(email=email)
        asunto = 'Gracias por suscribirte'
        mensaje = 'Te has suscripto exitosamente. Muchas gracias'
        # EMAIL_HOST_USER (cambiar en local)
        try:
            send_mail(asunto, mensaje, EMAIL_HOST_USER, [email])
        except:
            pass
        return redirect('index')


@method_decorator(login_required, name='dispatch')
class SuscritorAdminView(FilterView):
    model = Suscriptor
    filterset_class = SuscriptorListFilter
    context_object_name = 'suscritors'
    template_name = 'suscriptor/suscriptorAdmin.html'
    paginate_by = 30  # TODO obtener dato de un constants.py
    ordering = ["-date"]

    def get_queryset(self):
        return Suscriptor.objects.all()


@method_decorator(login_required, name='dispatch')
class ExcelSuscritorAdminView(TemplateView):
    def get(self, request, *args, **kwargs):
        subs = Suscriptor.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Lista de suscriptores'
        time = timezone.now()

        ws.merge_cells('B1:E1')

        ws['B3'] = 'ID'
        ws['C3'] = 'Nombre'
        ws['D3'] = 'E-mail'
        ws['E3'] = 'Estado'

        cont = 4

        for sub in subs:
            if sub.active == True:
                state = "activo"
            else:
                state = "inactivo"

            ws.cell(row=cont, column=2).value = sub.id
            ws.cell(row=cont, column=3).value = sub.name
            ws.cell(row=cont, column=4).value = sub.email
            ws.cell(row=cont, column=5).value = state

            cont += 1

        filename = str(time) + "subscritores.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(filename)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class SuscritorUpdateView(UpdateView):
    model = Suscriptor
    form_class = SuscriptorEmailForm
    template_name_suffix = '_update_form'
    template_name = 'suscriptor/suscriptorUpdate.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('suscriptoradminlist')

    def get_form(self, form_class=None):
        form = super(SuscritorUpdateView, self).get_form()
        # Modificar en tiempo real
        form.fields['email'].widget = forms.EmailInput(
            attrs={'class': 'form-control mb-2', 'placeholder': 'Email'})
        return form


@method_decorator(login_required, name='dispatch')
class SuscritorDeleteView(DeleteView):
    # Vista del contendido de cada post
    model = Suscriptor
    success_url = reverse_lazy('suscriptoradminlist')
    template_name = 'suscriptor/suscriptor_confirm_delete.html'


@method_decorator(login_required, name='dispatch')
class SuscritorSearchView(View):
    def post(self, request, *args, **kwargs):
        queryset = request.POST.get("buscar")
        if queryset:
            posts = Suscriptor.objects.filter(
                Q(email__icontains=queryset) |
                Q(name__icontains=queryset)
            ).distinct().order_by('-date')

        context = {
            'object_list': posts,
        }

        return render(request, 'suscriptor/suscriptor_result.html', context)


@method_decorator(login_required, name='dispatch')
class SuscriptorChageStateView(View):
    # Cambia el estado de un blog, usado en el jquerry de blogadmin-list.html
    def post(self, request, subcritor_id, *args, **kwargs):
        try:
            subcritor = Suscriptor.objects.get(id=subcritor_id)
            subcritor.active = not subcritor.active
            subcritor.save()
            if subcritor.active:
                audit_description = f"Se activa el Sitio :{subcritor_id}"
            else:
                audit_description = f"Se desactiva el Sitio :{subcritor_id}"
            Audits.audit_action(request.user, "", audit_description, "DELETE")
            return HttpResponse("ok", status=200)
        except Suscriptor.DoesNotExist:
            return HttpResponse("Cita not found", status=404)
        except Exception as e:
            print(e)
            return HttpResponse("error code", status=500)


@method_decorator(login_required, name='dispatch')
class FriendSitesCreateView(CreateView):
    # Vista para crear los post con un de decorador para que solo los administardores puedan acceder
    model = FriendSites
    form_class = FriendSitesForm
    success_url = reverse_lazy('administration')
    template_name = 'othersides/friendssites_form.html'

    # Devuelve al usuario actualmente logueado para el campo de User del Blog creado
    def form_valid(self, form):
        form.instance.user = self.request.user
        return super(FriendSitesCreateView, self).form_valid(form)


@method_decorator(login_required, name='dispatch')
class FriendSitesUpdateView(UpdateView):
    model = FriendSites
    form_class = FriendSitesForm
    template_name = 'othersides/friendssites_update_form.html'
    template_name_suffix = '_update_form'

    def get_success_url(self):
        return reverse_lazy('friendSiteUpdate', args=[self.object.id]) + '?ok'


@method_decorator(login_required, name='dispatch')
class FriendSitesAdminView(FilterView):
    model = FriendSites
    filterset_class = FriendSitesFilter
    context_object_name = 'friendsites'
    template_name = 'othersides/friendssitesAdmin.html'
    paginate_by = 30  # TODO obtener dato de un constants.py
    ordering = ["date"]

    def get_queryset(self):
        return FriendSites.objects.all()


@method_decorator(login_required, name='dispatch')
class FriendSitesChageStateView(View):
    # Cambia el estado de un blog, usado en el jquerry de blogadmin-list.html
    def post(self, request, friendsite_id, *args, **kwargs):
        try:
            friendsite = FriendSites.objects.get(id=friendsite_id)
            friendsite.active = not friendsite.active
            friendsite.save()
            if friendsite.active:
                audit_description = f"Se activa el Sitio :{friendsite_id}"
            else:
                audit_description = f"Se desactiva el Sitio :{friendsite_id}"
            Audits.audit_action(request.user, "", audit_description, "DELETE")
            return HttpResponse("ok", status=200)
        except FriendSites.DoesNotExist:
            return HttpResponse("Cita not found", status=404)
        except Exception as e:
            print(e)
            return HttpResponse("error code", status=500)
